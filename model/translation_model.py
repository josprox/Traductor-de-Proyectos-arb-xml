import os
import json
import requests
import csv
from datetime import datetime
import subprocess
import re


import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils import get_column_letter
from lxml import etree
import concurrent.futures
import urllib.parse

PLACEHOLDER_REGEX = re.compile(r'(%[0-9]+\$[0-9]*\.?[0-9]*[dsfpxXgGeE@]|%[0-9]*\.?[0-9]*[dsfpxXgGeE@]|\{[a-zA-Z0-9_]+\})')
RESTORE_REGEX = re.compile(r'___\s*[pP][hH]\s*_\s*(\d+)\s*___')


class TranslationCore:
    """
    Clase que encapsula toda la lógica de negocio de la aplicación:
    manipulación de archivos de idioma (ARB/XML), llamadas a la API de traducción,
    gestión del log y del historial.
    """
    FLUTTER_LANGUAGE_FILES = [
        "intl_en.arb", "intl_ar.arb", "intl_be.arb", "intl_bg.arb", "intl_bn.arb", "intl_cs.arb", "intl_de.arb",
        "intl_el.arb", "intl_es.arb", "intl_fa.arb", "intl_fi.arb", "intl_fr.arb",
        "intl_hu.arb", "intl_id.arb", "intl_it.arb", "intl_ja.arb", "intl_ko.arb", "intl_ml.arb",
        "intl_nb.arb", "intl_nl.arb", "intl_or.arb", "intl_pa.arb", "intl_pl.arb", "intl_pt.arb",
        "intl_ru.arb", "intl_sv.arb", "intl_tr.arb", "intl_uk.arb", "intl_vi.arb", "intl_zh.arb"
    ]

    KOTLIN_LANGUAGE_FOLDERS = [
        "values", "values-ar", "values-be", "values-bg", "values-bn", "values-bn-rIN", "values-bs", "values-cs", "values-de",
        "values-el", "values-es", "values-et", "values-fa", "values-fi", "values-fr", "values-hi", "values-hr",
        "values-hu", "values-in", "values-it", "values-ja", "values-ko", "values-ml",
        "values-nb-rNO", "values-ne", "values-nl", "values-or", "values-pa", "values-pl", "values-pt", "values-pt-rBR",
        "values-ru", "values-sv", "values-ta", "values-tr", "values-uk", "values-vi", "values-zh-rCN",
        "values-zh-rTW"
    ]
    KOTLIN_STRINGS_FILE_NAME = "strings-joss.xml"

    # API_URL eliminada, se usa lógica local con Google Translate
    GOOGLE_TRANSLATE_URL = "https://translate.googleapis.com/translate_a/single"
    LOG_FILE = "translation_log.xlsx"
    HISTORY_FILE = "translation_history.json"

    def __init__(self, project_path, log_callback=None):
        """
        Inicializa la lógica central de la aplicación.
        :param project_path: La ruta base del proyecto donde se encuentran los archivos de idioma.
        :param log_callback: Una función de callback para enviar mensajes de log a la UI.
        """
        self.project_path = project_path
        self.log_callback = log_callback if log_callback else print # Usa print si no se proporciona callback
        self.history = []
        self._load_history()
        self._initialize_log_file()

    def set_project_path(self, new_path):
        """Actualiza la ruta del proyecto y recarga historial/log."""
        self.project_path = new_path
        self._load_history()
        self._initialize_log_file()
        self._log(f"Ruta del proyecto actualizada a: {self.project_path}")

    def _log(self, message):
        """Envía un mensaje a la función de log configurada."""
        self.log_callback(message)

    def _initialize_log_file(self):
        """
        Asegura que el archivo de log XLSX exista y tenga los encabezados correctos.
        """
        log_path = os.path.join(self.project_path, self.LOG_FILE)
        if not os.path.exists(log_path):
            try:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Registro de Traducciones"
                headers = ["Fecha", "Idioma Base", "Texto Original", "Etiqueta", "Descripción", "Acción", "Plataforma"]
                sheet.append(headers)
                for i, header in enumerate(headers):
                    sheet.column_dimensions[get_column_letter(i + 1)].width = 20
                workbook.save(log_path)
                self._log(f"📝 Archivo de log '{self.LOG_FILE}' creado en {self.project_path}.")
            except Exception as e:
                self._log(f"❌ Error al crear el archivo de log XLSX en {self.project_path}: {e}")

    def _log_action(self, base_lang, original_text, key, desc, action_type, platform):
        """
        Registra una acción en el archivo XLSX.
        """
        log_path = os.path.join(self.project_path, self.LOG_FILE)
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            workbook = openpyxl.load_workbook(log_path)
            sheet = workbook.active
            sheet.append([timestamp, base_lang, original_text, key, desc, action_type, platform])
            workbook.save(log_path)
            self._log(f"📝 Acción '{action_type}' logeada para '{key}' en {platform.upper()}.")
        except Exception as e:
            self._log(f"❌ Error al escribir en el archivo de log XLSX: {e}")

    def _load_history(self):
        """
        Carga el historial de acciones desde el archivo JSON.
        """
        history_path = os.path.join(self.project_path, self.HISTORY_FILE)
        if os.path.exists(history_path):
            try:
                with open(history_path, 'r', encoding='utf-8') as f:
                    self.history = json.load(f)
                self._log(f"📚 Historial cargado desde '{self.HISTORY_FILE}' en {self.project_path}.")
            except json.JSONDecodeError:
                self._log(f"⚠️ Error al leer el historial. El archivo '{self.HISTORY_FILE}' puede estar corrupto.")
                self.history = []
            except Exception as e:
                self._log(f"❌ Error al cargar el historial: {e}")
                self.history = []
        else:
            self._log(f"ℹ️ No se encontró el archivo de historial '{self.HISTORY_FILE}' en {self.project_path}. Se creará uno nuevo.")

    def _save_history(self):
        """
        Guarda el historial de acciones actual en el archivo JSON.
        """
        history_path = os.path.join(self.project_path, self.HISTORY_FILE)
        try:
            with open(history_path, 'w', encoding='utf-8') as f:
                json.dump(self.history, f, indent=2, ensure_ascii=False)
            self._log(f"💾 Historial guardado en '{self.HISTORY_FILE}' en {self.project_path}.")
        except Exception as e:
            self._log(f"❌ Error al guardar el historial: {e}")

    def add_to_history(self, action_type, data, platform):
        """
        Añade una acción a la lista de historial y la guarda.
        """
        self.history.append({'type': action_type, 'data': data, 'timestamp': datetime.now().isoformat(), 'platform': platform})
        self._save_history()

    def get_history(self):
        """Devuelve el historial actual."""
        return self.history

    def pop_last_history_entry(self):
        """Elimina y devuelve la última entrada del historial."""
        if self.history:
            last_entry = self.history.pop()
            self._save_history() # Guardar el historial después de eliminar
            return last_entry
        return None

    @staticmethod
    def get_flutter_target_languages():
        """
        Extrae los códigos de idioma objetivo de la lista de archivos ARB.
        """
        return [file.split('_')[1].split('.')[0] for file in TranslationCore.FLUTTER_LANGUAGE_FILES]

    @staticmethod
    def get_kotlin_target_languages_for_api():
        """
        Construye una lista de códigos de idioma para la API a partir de KOTLIN_LANGUAGE_FOLDERS.
        Adapta los nombres de carpetas a códigos de idioma simples para la API.
        """
        target_langs = []
        for folder in TranslationCore.KOTLIN_LANGUAGE_FOLDERS:
            if folder == "values":
                target_langs.append("en") # 'values' suele ser el idioma por defecto, a menudo inglés
            else:
                lang_code = folder.replace("values-", "")
                # Manejar códigos regionales si la API espera solo la parte del idioma principal
                # Por ejemplo, 'bn-rIN' -> 'bn'
                if '-r' in lang_code:
                    lang_code = lang_code.split('-r')[0]
                target_langs.append(lang_code)
        return target_langs

    def create_flutter_language_files(self):
        """
        Crea nuevos archivos ARB para cada idioma si no existen.
        """
        self._log("Iniciando creación de archivos ARB (Flutter)...")
        for file_name in self.FLUTTER_LANGUAGE_FILES:
            path = os.path.join(self.project_path, file_name)
            locale = file_name.split('_')[1].split('.')[0]
            if not os.path.exists(path):
                try:
                    with open(path, "w", encoding="utf-8") as f:
                        json.dump({"@@locale": locale}, f, indent=2, ensure_ascii=False)
                    self._log(f"✅ Archivo creado: {file_name}")
                except Exception as e:
                    self._log(f"❌ Error al crear '{file_name}': {e}")
            else:
                self._log(f"⚠️ Archivo ya existe: {file_name}")
        self._log("Creación de archivos ARB (Flutter) finalizada.")

    def create_kotlin_language_folders(self):
        """
        Crea las carpetas y archivos strings-joss.xml para los idiomas de Kotlin si no existen.
        """
        self._log("Iniciando creación de carpetas y archivos XML (Kotlin)...")
        for folder_name in self.KOTLIN_LANGUAGE_FOLDERS:
            folder_path = os.path.join(self.project_path, folder_name)
            strings_file = os.path.join(folder_path, self.KOTLIN_STRINGS_FILE_NAME)

            if not os.path.exists(folder_path):
                try:
                    os.makedirs(folder_path)
                    self._log(f"✅ Carpeta creada: {folder_path}")
                except Exception as e:
                    self._log(f"❌ Error al crear carpeta '{folder_path}': {e}")
            else:
                self._log(f"⚠️ Carpeta ya existe: {folder_path}")

            if not os.path.exists(strings_file):
                try:
                    with open(strings_file, "w", encoding="utf-8") as file:
                        file.write("<?xml version='1.0' encoding='UTF-8'?>\n<resources>\n\n</resources>")
                    self._log(f"✅ Archivo creado: {strings_file}")
                except Exception as e:
                    self._log(f"❌ Error al crear archivo '{strings_file}': {e}")
            else:
                self._log(f"⚠️ Archivo ya existente: {strings_file}")
        self._log("Creación de carpetas y archivos XML (Kotlin) finalizada.")

    def delete_flutter_language_files(self):
        """
        Elimina todos los archivos ARB de Flutter.
        """
        self._log("Iniciando eliminación de archivos ARB (Flutter)...")
        for file_name in self.FLUTTER_LANGUAGE_FILES:
            path = os.path.join(self.project_path, file_name)
            if os.path.exists(path):
                try:
                    os.remove(path)
                    self._log(f"🗑️ Archivo eliminado: {file_name}")
                except Exception as e:
                    self._log(f"❌ Error al eliminar '{file_name}': {e}")
            else:
                self._log(f"⚠️ Archivo no encontrado: {file_name}")
        self._log("Eliminación de archivos ARB (Flutter) finalizada.")

    def delete_kotlin_language_folders(self):
        """
        Elimina las carpetas de idioma de Kotlin y sus contenidos.
        """
        self._log("Iniciando eliminación de carpetas y archivos XML (Kotlin)...")
        for folder_name in self.KOTLIN_LANGUAGE_FOLDERS:
            folder_path = os.path.join(self.project_path, folder_name)
            if os.path.exists(folder_path) and os.path.isdir(folder_path):
                try:
                    for root, dirs, files in os.walk(folder_path, topdown=False):
                        for file in files:
                            os.remove(os.path.join(root, file))
                        for dir in dirs:
                            os.rmdir(os.path.join(root, dir))
                    os.rmdir(folder_path)
                    self._log(f"🗑️ Carpeta eliminada: {folder_path}")
                except Exception as e:
                    self._log(f"❌ Error al eliminar '{folder_path}': {e}")
            else:
                self._log(f"⚠️ Carpeta no encontrada o ya eliminada: {folder_path}")
        self._log("Eliminación de carpetas y archivos XML (Kotlin) finalizada.")

    def check_key_existence(self, key, platform):
        """
        Verifica si una clave/string ya existe en los archivos de idioma para la plataforma dada.
        Retorna una lista de los nombres de archivos/carpetas donde la clave ya existe.
        """
        existing_key_locations = []
        if platform == "flutter":
            files_to_check = self.FLUTTER_LANGUAGE_FILES
        elif platform == "kotlin":
            files_to_check = [os.path.join(f, self.KOTLIN_STRINGS_FILE_NAME) for f in self.KOTLIN_LANGUAGE_FOLDERS]
        else:
            self._log("❌ Plataforma no válida para la verificación de clave.")
            return []

        for file_or_folder_path_part in files_to_check:
            full_path = os.path.join(self.project_path, file_or_folder_path_part)
            
            if platform == "flutter":
                if os.path.exists(full_path):
                    try:
                        with open(full_path, "r", encoding="utf-8") as f:
                            arb_data = json.load(f)
                        if key in arb_data:
                            existing_key_locations.append(file_or_folder_path_part)
                    except json.JSONDecodeError:
                        self._log(f"⚠️ Error al leer '{full_path}'. Archivo JSON inválido.")
                    except Exception as e:
                        self._log(f"❌ Error al verificar '{full_path}': {e}")
            elif platform == "kotlin":
                if os.path.exists(full_path):
                    try:
                        parser = etree.XMLParser(remove_blank_text=True, remove_comments=False)
                        tree = etree.parse(full_path, parser)
                        root = tree.getroot()
                        if any(child.get("name") == key for child in root.xpath("string")):
                            existing_key_locations.append(os.path.dirname(file_or_folder_path_part)) 
                    except etree.XMLSyntaxError:
                        self._log(f"⚠️ Error al leer '{full_path}'. Archivo XML inválido.")
                    except Exception as e:
                        self._log(f"❌ Error al verificar '{full_path}': {e}")
        return existing_key_locations

    def protect_placeholders(self, text):
        if not text:
            return text, []
        placeholders = []
        def replace_match(match):
            ph = match.group(0)
            placeholders.append(ph)
            return f"___PH_{len(placeholders) - 1}___"
        
        protected_text = PLACEHOLDER_REGEX.sub(replace_match, text)
        return protected_text, placeholders

    def restore_placeholders(self, text, placeholders):
        if not text or not placeholders:
            return text
        
        def restore_match(match):
            index = int(match.group(1))
            if index < len(placeholders):
                return placeholders[index]
            return match.group(0)
            
        return RESTORE_REGEX.sub(restore_match, text)

    def fetch_translations_from_api(self, base_lang, original_text, platform):
        """
        Realiza llamadas concurrentes a la API de Google Translate (gtx) para obtener traducciones.
        Simula la respuesta que antes daba la API de Laravel.
        """
        if platform == "flutter":
            target_langs = self.get_flutter_target_languages()
        elif platform == "kotlin":
            target_langs = self.get_kotlin_target_languages_for_api()
        else:
            raise ValueError("Plataforma no reconocida para la traducción.")

        translations = {}
        
        # Mapeo de idiomas para la API de Google (similares a los de TranslationController.php)
        # La mayoría coinciden, pero aseguramos la compatibilidad
        LANGUAGE_MAP = {
            'ar': 'ar', 'be': 'be', 'bg': 'bg', 'bn': 'bn', 'bs': 'bs', 'cs': 'cs', 'de': 'de',
            'el': 'el', 'en': 'en', 'es': 'es', 'et': 'et', 'fa': 'fa', 'fi': 'fi', 'fr': 'fr',
            'hi': 'hi', 'hr': 'hr', 'hu': 'hu', 'id': 'id', 'it': 'it', 'ja': 'ja', 'ko': 'ko',
            'ml': 'ml', 'nb': 'no', 'ne': 'ne', 'nl': 'nl', 'or': 'or', 'pa': 'pa', 'pl': 'pl',
            'pt': 'pt', 'ru': 'ru', 'sv': 'sv', 'ta': 'ta', 'tr': 'tr', 'uk': 'uk', 'vi': 'vi',
            'zh': 'zh-CN'
        }

        # Proteger placeholders antes de enviar a traducir
        protected_text, placeholders = self.protect_placeholders(original_text)

        # Función auxiliar para traducir un solo idioma
        def translate_single(target_code):
            google_target = LANGUAGE_MAP.get(target_code, target_code)
            
            # Casos especiales de mapeo (según el controlador original)
            if target_code == 'in': google_target = 'id'
            if target_code == 'nb-rNO': google_target = 'no'
            # Otros mapeos regionales suelen simplificarse por Google, pero enviamos el base si es simple
            
            try:
                params = {
                    'client': 'gtx',
                    'sl': base_lang,
                    'tl': google_target,
                    'dt': 't',
                    'q': protected_text
                }
                
                # Usar requests directamente
                response = requests.get(self.GOOGLE_TRANSLATE_URL, params=params, timeout=10)
                response.raise_for_status()
                
                # La respuesta es un array anidado: [[[ "Texto Traducido", "Texto Orig", ...], ...], ...]
                data = response.json()
                if data and len(data) > 0 and len(data[0]) > 0 and len(data[0][0]) > 0:
                    translated = data[0][0][0]
                    restored = self.restore_placeholders(translated, placeholders)
                    return target_code, restored
                else:
                    return target_code, None
            except Exception as e:
                self._log(f"⚠️ Error traduciendo a {target_code}: {e}")
                return target_code, None

        # Ejecución paralela
        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
            future_to_lang = {executor.submit(translate_single, lang): lang for lang in target_langs}
            for future in concurrent.futures.as_completed(future_to_lang):
                lang_code = future_to_lang[future]
                try:
                    code, text = future.result()
                    if text:
                        translations[code] = text
                    else:
                        translations[code] = original_text # Fallback al original si falla
                except Exception as exc:
                    self._log(f"Completado con excepción para {lang_code}: {exc}")
                    translations[lang_code] = original_text

        return translations

    def add_translation_entry(self, base_lang, original_text, key, desc, translations, existing_key_files, platform):
        """
        Añade una entrada de traducción a los archivos de idioma.
        Retorna los datos necesarios para la operación de deshacer.
        """
        undo_data = {
            'base_lang': base_lang,
            'original_text': original_text,
            'key': key,
            'desc': desc,
            'affected_files': {}
        }

        if platform == "flutter":
            target_assets = self.FLUTTER_LANGUAGE_FILES
        elif platform == "kotlin":
            target_assets = self.KOTLIN_LANGUAGE_FOLDERS
        else:
            raise ValueError(f"Plataforma desconocida: {platform}")

        for asset_name in target_assets:
            if platform == "flutter":
                current_path = os.path.join(self.project_path, asset_name)
                lang = asset_name.split('_')[1].split('.')[0]
                asset_identifier_for_check = asset_name 
            elif platform == "kotlin":
                current_path = os.path.join(self.project_path, asset_name, self.KOTLIN_STRINGS_FILE_NAME)
                lang = "en" if asset_name == "values" else asset_name.replace("values-", "")
                asset_identifier_for_check = asset_name 
            
            text = original_text if lang == base_lang else translations.get(lang)
            if not text and '-r' in lang:
                simple_lang = lang.split('-r')[0]
                text = translations.get(simple_lang)
                if text:
                    self._log(f"ℹ️ Usando traducción de '{simple_lang}' para '{lang}'.")
                else:
                    self._log(f"⚠️ Traducción omitida para {lang} (texto no disponible).")
                    continue
            elif not text:
                self._log(f"⚠️ Traducción omitida para {lang} (texto no disponible).")
                continue

            if asset_identifier_for_check in existing_key_files:
                self._log(f"ℹ️ Clave/string '{key}' ya existe en '{asset_identifier_for_check}'. Se omite la adición para esta ubicación.")
                continue

            if not os.path.exists(os.path.dirname(current_path)): # Asegurarse de que la carpeta exista
                os.makedirs(os.path.dirname(current_path), exist_ok=True)

            if not os.path.exists(current_path):
                self._log(f"❌ Archivo/ubicación no encontrado: {current_path}. Creando...")
                if platform == "flutter":
                    with open(current_path, "w", encoding="utf-8") as f:
                        json.dump({"@@locale": lang}, f, indent=2, ensure_ascii=False)
                elif platform == "kotlin":
                    with open(current_path, "w", encoding="utf-8") as f:
                        f.write("<?xml version='1.0' encoding='UTF-8'?>\n<resources>\n\n</resources>")

            try:
                if platform == "flutter":
                    with open(current_path, "r", encoding="utf-8") as f:
                        arb_data = json.load(f)
                    
                    undo_data['affected_files'][asset_name] = {
                        'old_value': arb_data.get(key),
                        'old_desc': arb_data.get(f"@{key}", {}).get("description")
                    }

                    arb_data[key] = text
                    arb_data[f"@{key}"] = {"description": desc}

                    with open(current_path, "w", encoding="utf-8") as f:
                        json.dump(arb_data, f, indent=2, ensure_ascii=False)
                    self._log(f"✅ Añadido '{key}' en {asset_name} (Flutter)")

                elif platform == "kotlin":
                    parser = etree.XMLParser(remove_blank_text=True, remove_comments=False)
                    tree = etree.parse(current_path, parser)
                    root = tree.getroot()

                    existing_string_element = root.xpath(f"string[@name='{key}']")
                    if existing_string_element:
                        undo_data['affected_files'][asset_name] = {
                            'old_value': existing_string_element[0].text,
                            'old_desc': None
                        }
                    else:
                        undo_data['affected_files'][asset_name] = {'old_value': None, 'old_desc': None}

                    if not existing_string_element:
                        new_string = etree.Element("string", name=key)
                        new_string.text = text
                        root.append(new_string)

                        formatted_xml = etree.tostring(
                            root,
                            encoding="utf-8",
                            xml_declaration=True,
                            pretty_print=True
                        ).decode("utf-8")
                        # Asegurar una nueva línea final si pretty_print no la añade consistentemente
                        if not formatted_xml.endswith("\n"):
                            formatted_xml += "\n"
                        
                        with open(current_path, "w", encoding="utf-8") as f:
                            f.write(formatted_xml)
                        self._log(f"✅ Añadido '{key}' en {asset_name}/{self.KOTLIN_STRINGS_FILE_NAME} (Kotlin)")
                    else:
                        self._log(f"⚠️ String '{key}' ya existe en {asset_name}/{self.KOTLIN_STRINGS_FILE_NAME}. Se omite la adición.")

            except json.JSONDecodeError:
                self._log(f"⚠️ Error al leer '{current_path}'. Archivo JSON/XML inválido.")
            except Exception as e:
                self._log(f"❌ Error al procesar '{current_path}': {e}")
        
        self._log_action(base_lang, original_text, key, desc, "Traducción y Adición", platform)
        self.add_to_history('add_key', undo_data, platform)
        self._log(f"Traducción y adición de etiqueta/string finalizada para {platform.upper()}.")
        return undo_data # Retorna los datos para el historial

    def delete_key_entry(self, key, platform):
        """
        Elimina una clave/string de los archivos de idioma.
        Retorna los datos necesarios para la operación de deshacer.
        """
        undo_data = {'key': key, 'deleted_content_per_file': {}}
        
        if platform == "flutter":
            assets_to_check = self.FLUTTER_LANGUAGE_FILES
        elif platform == "kotlin":
            assets_to_check = self.KOTLIN_LANGUAGE_FOLDERS
        else:
            raise ValueError("Plataforma no válida para la eliminación de clave/string.")

        for asset_name in assets_to_check:
            if platform == "flutter":
                current_path = os.path.join(self.project_path, asset_name)
            elif platform == "kotlin":
                current_path = os.path.join(self.project_path, asset_name, self.KOTLIN_STRINGS_FILE_NAME)

            if os.path.exists(current_path):
                try:
                    if platform == "flutter":
                        with open(current_path, "r", encoding="utf-8") as f:
                            arb_data = json.load(f)
                        if key in arb_data:
                            undo_data['deleted_content_per_file'][asset_name] = {
                                'value': arb_data.get(key),
                                'description': arb_data.get(f"@{key}", {}).get("description")
                            }
                            del arb_data[key]
                            arb_data.pop(f"@{key}", None)
                            with open(current_path, "w", encoding="utf-8") as f:
                                json.dump(arb_data, f, indent=2, ensure_ascii=False)
                            self._log(f"🗑️ '{key}' eliminado de {asset_name} (Flutter)")
                        else:
                            self._log(f"⚠️ '{key}' no encontrado en {asset_name} (Flutter)")
                    elif platform == "kotlin":
                        parser = etree.XMLParser(remove_blank_text=True, remove_comments=False)
                        tree = etree.parse(current_path, parser)
                        root = tree.getroot()

                        target_string = root.xpath(f"string[@name='{key}']")
                        if target_string:
                            undo_data['deleted_content_per_file'][asset_name] = {
                                'value': target_string[0].text,
                                'description': None
                            }
                            root.remove(target_string[0])
                            formatted_xml = etree.tostring(
                                root,
                                encoding="utf-8",
                                xml_declaration=True,
                                pretty_print=True
                            ).decode("utf-8")
                            # Asegurar una nueva línea final si pretty_print no la añade consistentemente
                            if not formatted_xml.endswith("\n"):
                                formatted_xml += "\n"
                            with open(current_path, "w", encoding="utf-8") as f:
                                f.write(formatted_xml)
                            self._log(f"🗑️ '{key}' eliminado de {asset_name}/{self.KOTLIN_STRINGS_FILE_NAME} (Kotlin)")
                        else:
                            self._log(f"⚠️ '{key}' no encontrado en {asset_name}/{self.KOTLIN_STRINGS_FILE_NAME} (Kotlin)")

                except (json.JSONDecodeError, etree.XMLSyntaxError):
                    self._log(f"⚠️ Error al leer '{current_path}'. Archivo inválido. No se puede guardar para deshacer.")
                except Exception as e:
                    self._log(f"❌ Error al acceder a '{current_path}': {e}. No se puede guardar para deshacer.")
            else:
                self._log(f"⚠️ Archivo/ubicación no encontrado: {current_path}")
        
        self._log_action("", "", key, "", "Eliminación de Etiqueta/String", platform)
        self.add_to_history('delete_key', undo_data, platform)
        self._log(f"Eliminación de clave/string '{key}' finalizada para {platform.upper()}.")
        return undo_data

    def run_flutter_intl_generate(self):
        """
        Ejecuta el comando 'dart run intl_utils:generate' en el directorio del proyecto.
        Retorna la salida del comando y el código de retorno.
        """
        if not self.project_path:
            raise ValueError("No se ha seleccionado una carpeta de proyecto para ejecutar el comando.")

        self._log(f"Ejecutando 'dart run intl_utils:generate' en: {self.project_path}")

        try:
            command = ["dart", "run", "intl_utils:generate"]
            process = subprocess.run(
                command,
                cwd=self.project_path,
                capture_output=True,
                text=True,
                check=False
            )
            output = f"--- Salida del comando ---\n{process.stdout}\n--- Errores del comando ---\n{process.stderr}"
            if process.returncode == 0:
                self._log("✅ Comando 'dart run intl_utils:generate' ejecutado con éxito.")
            else:
                self._log(f"❌ El comando 'dart run intl_utils:generate' falló con código de salida {process.returncode}.")
            return output, process.returncode
        except FileNotFoundError:
            raise FileNotFoundError("Error: 'dart' o 'intl_utils' no encontrado. Asegúrate de que Flutter SDK esté en tu PATH y que intl_utils esté configurado en tu proyecto.")
        except Exception as e:
            raise Exception(f"Error al ejecutar el comando: {e}")

    def undo_delete_key_action(self, key_to_delete, platform):
        """
        Método interno para eliminar una clave/string de todos los archivos/ubicaciones,
        usado para deshacer una acción 'add_key'.
        """
        if platform == "flutter":
            target_assets = self.FLUTTER_LANGUAGE_FILES
        elif platform == "kotlin":
            target_assets = self.KOTLIN_LANGUAGE_FOLDERS
        else:
            self._log(f"❌ Plataforma desconocida para deshacer eliminación: {platform}")
            return

        for asset_name in target_assets:
            if platform == "flutter":
                current_path = os.path.join(self.project_path, asset_name)
            elif platform == "kotlin":
                current_path = os.path.join(self.project_path, asset_name, self.KOTLIN_STRINGS_FILE_NAME)

            if os.path.exists(current_path):
                try:
                    if platform == "flutter":
                        with open(current_path, "r", encoding="utf-8") as f:
                            data = json.load(f)

                        if key_to_delete in data:
                            del data[key_to_delete]
                            data.pop(f"@{key_to_delete}", None)

                            with open(current_path, "w", encoding="utf-8") as f:
                                json.dump(data, f, indent=2, ensure_ascii=False)
                            self._log(f"🗑️ '{key_to_delete}' eliminado de {asset_name} (deshecho Flutter).")
                        else:
                            self._log(f"⚠️ '{key_to_delete}' no encontrado en {asset_name} (para deshacer Flutter).")
                    elif platform == "kotlin":
                        parser = etree.XMLParser(remove_blank_text=True, remove_comments=False)
                        tree = etree.parse(current_path, parser)
                        root = tree.getroot()

                        target_string = root.xpath(f"string[@name='{key_to_delete}']")
                        if target_string:
                            root.remove(target_string[0])
                            formatted_xml = etree.tostring(
                                root,
                                encoding="utf-8",
                                xml_declaration=True,
                                pretty_print=True
                            ).decode("utf-8")
                            # Asegurar una nueva línea final si pretty_print no la añade consistentemente
                            if not formatted_xml.endswith("\n"):
                                formatted_xml += "\n"
                            with open(current_path, "w", encoding="utf-8") as f:
                                f.write(formatted_xml)
                            self._log(f"🗑️ '{key_to_delete}' eliminado de {asset_name}/{self.KOTLIN_STRINGS_FILE_NAME} (deshecho Kotlin).")
                        else:
                            self._log(f"⚠️ '{key_to_delete}' no encontrado en {asset_name}/{self.KOTLIN_STRINGS_FILE_NAME} (para deshacer Kotlin).")

                except (json.JSONDecodeError, etree.XMLSyntaxError):
                    self._log(f"⚠️ Error al leer '{current_path}'. Archivo JSON/XML inválido (para deshacer).")
                except Exception as e:
                    self._log(f"❌ Error al procesar '{current_path}' (para deshacer): {e}")
            else:
                self._log(f"⚠️ Archivo/ubicación no encontrado: {current_path} (para deshacer).")

    def undo_add_key_action(self, key_to_restore, deleted_content_per_file, platform):
        """
        Método interno para volver a añadir una clave/string con su contenido anterior,
        usado para deshacer una acción 'delete_key'.
        """
        if platform == "flutter":
            target_assets = self.FLUTTER_LANGUAGE_FILES
        elif platform == "kotlin":
            target_assets = self.KOTLIN_LANGUAGE_FOLDERS
        else:
            self._log(f"❌ Plataforma desconocida para deshacer adición: {platform}")
            return

        for asset_name in target_assets:
            if platform == "flutter":
                current_path = os.path.join(self.project_path, asset_name)
            elif platform == "kotlin":
                current_path = os.path.join(self.project_path, asset_name, self.KOTLIN_STRINGS_FILE_NAME)

            content_to_restore = deleted_content_per_file.get(asset_name)

            if not content_to_restore or content_to_restore['value'] is None:
                self._log(f"⚠️ No hay contenido para restaurar '{key_to_restore}' en {asset_name}.")
                continue

            try:
                if platform == "flutter":
                    arb_data = {"@@locale": asset_name.split('_')[1].split('.')[0]}
                    if os.path.exists(current_path):
                        with open(current_path, "r", encoding="utf-8") as f:
                            arb_data = json.load(f)

                    arb_data[key_to_restore] = content_to_restore['value']
                    if content_to_restore['description']:
                        arb_data[f"@{key_to_restore}"] = {"description": content_to_restore['description']}
                    else:
                        arb_data.pop(f"@{key_to_restore}", None)

                    with open(current_path, "w", encoding="utf-8") as f:
                        json.dump(arb_data, f, indent=2, ensure_ascii=False)
                    self._log(f"✅ Restaurado '{key_to_restore}' en {asset_name} (deshecho Flutter).")

                elif platform == "kotlin":
                    parser = etree.XMLParser(remove_blank_text=True, remove_comments=False)
                    if not os.path.exists(current_path):
                        os.makedirs(os.path.dirname(current_path), exist_ok=True)
                        with open(current_path, "w", encoding="utf-8") as f:
                            f.write("<?xml version='1.0' encoding='UTF-8'?>\n<resources>\n\n</resources>")
                        tree = etree.parse(current_path, parser)
                        root = tree.getroot()
                    else:
                        tree = etree.parse(current_path, parser)
                        root = tree.getroot()

                    existing_string_element = root.xpath(f"string[@name='{key_to_restore}']")
                    if not existing_string_element:
                        new_string = etree.Element("string", name=key_to_restore)
                        new_string.text = content_to_restore['value']
                        root.append(new_string)

                        formatted_xml = etree.tostring(
                            root,
                            encoding="utf-8",
                            xml_declaration=True,
                            pretty_print=True
                        ).decode("utf-8")
                        # Asegurar una nueva línea final si pretty_print no la añade consistentemente
                        if not formatted_xml.endswith("\n"):
                            formatted_xml += "\n"
                        with open(current_path, "w", encoding="utf-8") as f:
                            f.write(formatted_xml)
                        self._log(f"✅ Restaurado '{key_to_restore}' en {asset_name}/{self.KOTLIN_STRINGS_FILE_NAME} (deshecho Kotlin).")
                    else:
                        self._log(f"⚠️ El string '{key_to_restore}' ya existe en {asset_name}/{self.KOTLIN_STRINGS_FILE_NAME}. No se restauró para evitar duplicados.")

            except (json.JSONDecodeError, etree.XMLSyntaxError):
                self._log(f"⚠️ Error al leer '{current_path}'. Archivo JSON/XML inválido (para deshacer).")
            except Exception as e:
                self._log(f"❌ Error al restaurar '{key_to_restore}' en '{current_path}' (deshacer): {e}")

    def parse_batch_content(self, content):
        """
        Detecta el formato (ARB JSON o Android XML) y lo analiza para extraer claves y valores.
        Retorna: (platform, base_lang, key_values, descriptions)
        """
        content_stripped = content.strip()
        if content_stripped.startswith("<") or "<resources" in content_stripped or "<string" in content_stripped:
            # Es XML de Android / Kotlin
            try:
                clean_content = content_stripped
                if "<?xml" in clean_content:
                    clean_content = re.sub(r'<\?xml.*?\?>', '', clean_content).strip()
                
                if not clean_content.startswith("<resources"):
                    clean_content = f"<resources>{clean_content}</resources>"
                
                parser = etree.XMLParser(remove_blank_text=True, remove_comments=False)
                root = etree.fromstring(clean_content.encode('utf-8'), parser)
                
                key_values = {}
                for child in root.xpath("string"):
                    name = child.get("name")
                    if name:
                        text = "".join(child.itertext()) if child.itertext() else (child.text or "")
                        key_values[name] = text
                
                return "kotlin", None, key_values, {}
            except Exception as e:
                raise ValueError(f"Error al parsear el XML de Android: {e}")
        else:
            # Es ARB JSON de Flutter
            try:
                data = json.loads(content_stripped)
                base_lang = data.get("@@locale")
                key_values = {}
                descriptions = {}
                for k, v in data.items():
                    if k.startswith("@@"):
                        continue
                    if k.startswith("@"):
                        real_key = k[1:]
                        if isinstance(v, dict) and "description" in v:
                            descriptions[real_key] = v["description"]
                    else:
                        if isinstance(v, str):
                            key_values[k] = v
                return "flutter", base_lang, key_values, descriptions
            except Exception as e:
                raise ValueError(f"Error al parsear el JSON de ARB: {e}")

    def add_translation_batch(self, base_lang, batch_translations, descriptions, platform):
        """
        Añade un lote de traducciones a los archivos correspondientes.
        Retorna los datos para el historial (deshacer).
        """
        undo_data = {
            'affected_files': {}
        }

        if platform == "flutter":
            target_assets = self.FLUTTER_LANGUAGE_FILES
        elif platform == "kotlin":
            target_assets = self.KOTLIN_LANGUAGE_FOLDERS
        else:
            raise ValueError(f"Plataforma desconocida: {platform}")

        for asset_name in target_assets:
            if platform == "flutter":
                current_path = os.path.join(self.project_path, asset_name)
                lang = asset_name.split('_')[1].split('.')[0]
            elif platform == "kotlin":
                current_path = os.path.join(self.project_path, asset_name, self.KOTLIN_STRINGS_FILE_NAME)
                lang = "en" if asset_name == "values" else asset_name.replace("values-", "")

            # Obtener traducciones para este idioma específico
            lang_translations = batch_translations.get(lang, {})
            if not lang_translations and '-r' in lang:
                simple_lang = lang.split('-r')[0]
                lang_translations = batch_translations.get(simple_lang, {})

            if not lang_translations:
                continue

            if not os.path.exists(os.path.dirname(current_path)):
                os.makedirs(os.path.dirname(current_path), exist_ok=True)

            if not os.path.exists(current_path):
                if platform == "flutter":
                    with open(current_path, "w", encoding="utf-8") as f:
                        json.dump({"@@locale": lang}, f, indent=2, ensure_ascii=False)
                elif platform == "kotlin":
                    with open(current_path, "w", encoding="utf-8") as f:
                        f.write("<?xml version='1.0' encoding='UTF-8'?>\n<resources>\n\n</resources>")

            try:
                undo_data['affected_files'][asset_name] = {}
                
                if platform == "flutter":
                    with open(current_path, "r", encoding="utf-8") as f:
                        arb_data = json.load(f)

                    for key, text in lang_translations.items():
                        # Guardar estado anterior para deshacer
                        undo_data['affected_files'][asset_name][key] = {
                            'old_value': arb_data.get(key),
                            'old_desc': arb_data.get(f"@{key}", {}).get("description")
                        }
                        
                        arb_data[key] = text
                        desc = descriptions.get(key)
                        if desc:
                            arb_data[f"@{key}"] = {"description": desc}

                    with open(current_path, "w", encoding="utf-8") as f:
                        json.dump(arb_data, f, indent=2, ensure_ascii=False)
                    self._log(f"✅ Lote de traducciones añadido en {asset_name} (Flutter)")

                elif platform == "kotlin":
                    parser = etree.XMLParser(remove_blank_text=True, remove_comments=False)
                    tree = etree.parse(current_path, parser)
                    root = tree.getroot()

                    for key, text in lang_translations.items():
                        existing_string_element = root.xpath(f"string[@name='{key}']")
                        
                        # Guardar estado anterior para deshacer
                        if existing_string_element:
                            undo_data['affected_files'][asset_name][key] = {
                                'old_value': existing_string_element[0].text,
                                'old_desc': None
                            }
                            existing_string_element[0].text = text
                        else:
                            undo_data['affected_files'][asset_name][key] = {
                                'old_value': None,
                                'old_desc': None
                            }
                            new_string = etree.Element("string", name=key)
                            new_string.text = text
                            root.append(new_string)

                    formatted_xml = etree.tostring(
                        root,
                        encoding="utf-8",
                        xml_declaration=True,
                        pretty_print=True
                    ).decode("utf-8")
                    if not formatted_xml.endswith("\n"):
                        formatted_xml += "\n"

                    with open(current_path, "w", encoding="utf-8") as f:
                        f.write(formatted_xml)
                    self._log(f"✅ Lote de traducciones añadido en {asset_name}/{self.KOTLIN_STRINGS_FILE_NAME} (Kotlin)")

            except Exception as e:
                self._log(f"❌ Error al procesar '{current_path}' en lote: {e}")

        # Se registra el lote con un texto genérico de resumen en el log
        num_keys = len(descriptions) if descriptions else len(list(batch_translations.values())[0]) if batch_translations else 0
        self._log_action(base_lang, f"[Lote de {num_keys} strings]", "Varios", "", "Traducción Lote", platform)
        self.add_to_history('batch_add_keys', undo_data, platform)
        return undo_data

    def undo_batch_add_keys_action(self, payload, platform):
        """
        Deshace un lote completo de traducciones.
        """
        affected_files = payload.get('affected_files', {})
        for asset_name, keys_data in affected_files.items():
            if platform == "flutter":
                current_path = os.path.join(self.project_path, asset_name)
            elif platform == "kotlin":
                current_path = os.path.join(self.project_path, asset_name, self.KOTLIN_STRINGS_FILE_NAME)
            else:
                continue

            if not os.path.exists(current_path):
                continue

            try:
                if platform == "flutter":
                    with open(current_path, "r", encoding="utf-8") as f:
                        data = json.load(f)

                    for key, values in keys_data.items():
                        old_val = values.get('old_value')
                        old_desc = values.get('old_desc')
                        if old_val is None:
                            data.pop(key, None)
                            data.pop(f"@{key}", None)
                        else:
                            data[key] = old_val
                            if old_desc:
                                data[f"@{key}"] = {"description": old_desc}
                            else:
                                data.pop(f"@{key}", None)

                    with open(current_path, "w", encoding="utf-8") as f:
                        json.dump(data, f, indent=2, ensure_ascii=False)
                    self._log(f"🗑️ Revertido lote en {asset_name} (Flutter).")

                elif platform == "kotlin":
                    parser = etree.XMLParser(remove_blank_text=True, remove_comments=False)
                    tree = etree.parse(current_path, parser)
                    root = tree.getroot()

                    for key, values in keys_data.items():
                        old_val = values.get('old_value')
                        target_string = root.xpath(f"string[@name='{key}']")
                        if old_val is None:
                            if target_string:
                                root.remove(target_string[0])
                        else:
                            if target_string:
                                target_string[0].text = old_val
                            else:
                                new_string = etree.Element("string", name=key)
                                new_string.text = old_val
                                root.append(new_string)

                    formatted_xml = etree.tostring(
                        root,
                        encoding="utf-8",
                        xml_declaration=True,
                        pretty_print=True
                    ).decode("utf-8")
                    if not formatted_xml.endswith("\n"):
                        formatted_xml += "\n"
                    with open(current_path, "w", encoding="utf-8") as f:
                        f.write(formatted_xml)
                    self._log(f"🗑️ Revertido lote en {asset_name}/{self.KOTLIN_STRINGS_FILE_NAME} (Kotlin).")

            except Exception as e:
                self._log(f"❌ Error al deshacer lote en '{asset_name}': {e}")
