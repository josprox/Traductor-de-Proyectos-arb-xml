import os
import sys
import json
import requests
import csv # Se mantiene por si se necesita para otras operaciones, aunque no para el log XLSX
from datetime import datetime
import subprocess # Para ejecutar comandos externos

# Importar openpyxl para manejar archivos .xlsx
import openpyxl
from openpyxl.utils import get_column_letter

# Importar lxml para manejar archivos XML (Kotlin)
from lxml import etree

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QPushButton, QWidget, QVBoxLayout,
    QLabel, QLineEdit, QTextEdit, QInputDialog, QMessageBox, QProgressBar,
    QHBoxLayout, QDialog, QListWidget, QListWidgetItem, QComboBox, QFileDialog
)
from PySide6.QtCore import Qt, QThread, Signal

# --- Constantes para Flutter (ARB) ---
FLUTTER_LANGUAGE_FILES = [
    "intl_en.arb", "intl_ar.arb", "intl_be.arb", "intl_bg.arb", "intl_bn.arb", "intl_cs.arb", "intl_de.arb",
    "intl_el.arb", "intl_es.arb", "intl_fa.arb", "intl_fi.arb", "intl_fr.arb",
    "intl_hu.arb", "intl_id.arb", "intl_it.arb", "intl_ja.arb", "intl_ko.arb", "intl_ml.arb",
    "intl_nb.arb", "intl_nl.arb", "intl_or.arb", "intl_pa.arb", "intl_pl.arb", "intl_pt.arb",
    "intl_ru.arb", "intl_sv.arb", "intl_tr.arb", "intl_uk.arb", "intl_vi.arb", "intl_zh.arb"
]

# --- Constantes para Kotlin (XML de Android) ---
KOTLIN_LANGUAGE_FOLDERS = [
    "values", "values-ar", "values-be", "values-bg", "values-bn", "values-bn-rIN", "values-bs", "values-cs", "values-de",
    "values-el", "values-es", "values-et", "values-fa", "values-fi", "values-fr", "values-hi", "values-hr",
    "values-hu", "values-in", "values-it", "values-ja", "values-ko", "values-ml",
    "values-nb-rNO", "values-ne", "values-nl", "values-or", "values-pa", "values-pl", "values-pt", "values-pt-rBR",
    "values-ru", "values-sv", "values-ta", "values-tr", "values-uk", "values-vi", "values-zh-rCN",
    "values-zh-rTW"
]
KOTLIN_STRINGS_FILE_NAME = "strings-joss.xml" # Nombre del archivo XML dentro de cada carpeta de idioma

API_URL = "https://jossred.josprox.com/api/traducir"
LOG_FILE = "translation_log.xlsx"
HISTORY_FILE = "translation_history.json"

def get_script_dir():
    """
    Devuelve el directorio donde se encuentra el script.
    """
    return os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.realpath(__file__))

def get_flutter_target_languages():
    """
    Extrae los códigos de idioma objetivo de la lista de archivos ARB.
    """
    return [file.split('_')[1].split('.')[0] for file in FLUTTER_LANGUAGE_FILES]

def get_kotlin_target_languages_for_api():
    """
    Construye una lista de códigos de idioma para la API a partir de KOTLIN_LANGUAGE_FOLDERS.
    Adapta los nombres de carpetas a códigos de idioma simples para la API.
    """
    target_langs = []
    for folder in KOTLIN_LANGUAGE_FOLDERS:
        if folder == "values":
            target_langs.append("en") # 'values' suele ser el idioma por defecto, a menudo inglés
        else:
            lang_code = folder.replace("values-", "")
            # Manejar códigos regionales si la API espera solo la parte del idioma principal
            # Por ejemplo, 'bn-rIN' -> 'bn'
            if '-r' in lang_code:
                lang_code = lang_code.split('-r')[0]
            target_langs.append(lang_code)
    return target_langs

class WorkerThread(QThread):
    """
    Una subclase de QThread para realizar operaciones de larga duración (como la traducción)
    en segundo plano, manteniendo la UI responsiva.
    """
    progress_updated = Signal(int)
    log_message = Signal(str)
    translation_finished = Signal(dict)
    error_occurred = Signal(str)
    command_output = Signal(str) # Nueva señal para la salida de comandos

    def __init__(self, operation_type, data=None, platform="flutter", project_path=None):
        super().__init__()
        self.operation_type = operation_type
        self.data = data
        self.platform = platform
        self.project_path = project_path # Ruta del proyecto para comandos

    def run(self):
        """
        Ejecuta la operación especificada según `operation_type` y la plataforma.
        """
        if self.operation_type == "translate_and_add":
            self._perform_translation_and_add()
        elif self.operation_type == "delete_key":
            # Aquí se llama a _perform_delete_key
            self._perform_delete_key()
        elif self.operation_type == "run_flutter_intl_generate":
            self._run_flutter_intl_generate_command()

    def _perform_translation_and_add(self):
        """
        Maneja la traducción y adición de una nueva clave a los archivos ARB/XML.
        Emite señales para el progreso, mensajes de log y finalización/errores.
        """
        base_lang = self.data['base_lang']
        original_text = self.data['original_text']
        key = self.data['key']
        desc = self.data['desc']
        existing_key_files = self.data.get('existing_key_files', [])

        self.log_message.emit(f"Iniciando traducción para '{key}' en plataforma {self.platform.upper()}...")
        self.progress_updated.emit(0)

        try:
            if self.platform == "flutter":
                langs_for_api = get_flutter_target_languages()
            elif self.platform == "kotlin":
                langs_for_api = get_kotlin_target_languages_for_api()
            else:
                self.error_occurred.emit("❌ Plataforma no reconocida para la traducción.")
                return

            response = requests.get(f"{API_URL}?idioma={base_lang}&texto={original_text}&traducir={','.join(langs_for_api)}")
            response.raise_for_status()
            data = response.json()
        except requests.exceptions.RequestException as e:
            self.error_occurred.emit(f"❌ Error de conexión con la API: {e}")
            return
        except json.JSONDecodeError:
            self.error_occurred.emit("❌ Error al decodificar la respuesta JSON de la API.")
            return

        if not data.get("success"):
            self.error_occurred.emit(f"❌ API Error: {data.get('error', 'Error desconocido')}")
            return

        translations = {t["target_language"]: t["translated_text"] for t in data.get("translations", [])}

        result_data = {
            'base_lang': base_lang,
            'original_text': original_text,
            'key': key,
            'desc': desc,
            'translations': translations,
            'existing_key_files': existing_key_files,
            'platform': self.platform
        }
        self.translation_finished.emit(result_data)

    def _run_flutter_intl_generate_command(self):
        """
        Ejecuta el comando 'dart run intl_utils:generate' en el directorio del proyecto.
        """
        if not self.project_path:
            self.error_occurred.emit("❌ No se ha seleccionado una carpeta de proyecto para ejecutar el comando.")
            return

        self.log_message.emit(f"Ejecutando 'dart run intl_utils:generate' en: {self.project_path}")
        self.progress_updated.emit(0) # Iniciar progreso

        try:
            command = ["dart", "run", "intl_utils:generate"]
            
            process = subprocess.run(
                command,
                cwd=self.project_path,
                capture_output=True,
                text=True,
                check=False
            )

            self.command_output.emit(f"--- Salida del comando ---\n{process.stdout}\n--- Errores del comando ---\n{process.stderr}")
            
            if process.returncode == 0:
                self.log_message.emit("✅ Comando 'dart run intl_utils:generate' ejecutado con éxito.")
            else:
                self.error_occurred.emit(f"❌ El comando 'dart run intl_utils:generate' falló con código de salida {process.returncode}.")

        except FileNotFoundError:
            self.error_occurred.emit("❌ Error: 'dart' o 'intl_utils' no encontrado. Asegúrate de que Flutter SDK esté en tu PATH y que intl_utils esté configurado en tu proyecto.")
        except Exception as e:
            self.error_occurred.emit(f"❌ Error al ejecutar el comando: {e}")
        finally:
            self.progress_updated.emit(100) # Finalizar progreso
            self.translation_finished.emit({'platform': self.platform}) # Señal de finalización para re-habilitar la UI

    # --- Método _perform_delete_key (Asegúrate de que este método esté correctamente indentado dentro de WorkerThread) ---
    def _perform_delete_key(self):
        """
        Método del hilo de trabajo para realizar la eliminación real de una clave/string.
        """
        key = self.data['key']
        undo_data = self.data['undo_data']
        platform = self.platform

        if platform == "flutter":
            target_assets = FLUTTER_LANGUAGE_FILES
        elif platform == "kotlin":
            target_assets = KOTLIN_LANGUAGE_FOLDERS
        else:
            self.log_message.emit(f"❌ Plataforma desconocida para la eliminación: {platform}")
            self.error_occurred.emit(f"Plataforma desconocida: {platform}")
            return

        files_processed_count = 0
        total_assets = len(target_assets)
        self.progress_updated.emit(0)

        for i, asset_name in enumerate(target_assets):
            if platform == "flutter":
                current_path = os.path.join(self.project_path, asset_name)
            elif platform == "kotlin":
                current_path = os.path.join(self.project_path, asset_name, KOTLIN_STRINGS_FILE_NAME)

            if os.path.exists(current_path):
                try:
                    if platform == "flutter":
                        with open(current_path, "r", encoding="utf-8") as f:
                            data = json.load(f)

                        if key in data:
                            del data[key]
                            data.pop(f"@{key}", None)

                            with open(current_path, "w", encoding="utf-8") as f:
                                json.dump(data, f, indent=2, ensure_ascii=False)
                            self.log_message.emit(f"🗑️ '{key}' eliminado de {asset_name} (Flutter)")
                        else:
                            self.log_message.emit(f"⚠️ '{key}' no encontrado en {asset_name} (Flutter)")
                    elif platform == "kotlin":
                        parser = etree.XMLParser(remove_blank_text=True, remove_comments=False)
                        tree = etree.parse(current_path, parser)
                        root = tree.getroot()

                        target_string = root.xpath(f"string[@name='{key}']")
                        if target_string:
                            root.remove(target_string[0])
                            formatted_xml = etree.tostring(
                                root,
                                encoding="utf-8",
                                xml_declaration=True,
                                pretty_print=True
                            ).decode("utf-8")
                            if not formatted_xml.endswith("\n</resources>\n"):
                                formatted_xml = formatted_xml.replace("</resources>", "\n</resources>")
                            with open(current_path, "w", encoding="utf-8") as f:
                                f.write(formatted_xml)
                            self.log_message.emit(f"🗑️ '{key}' eliminado de {asset_name}/{KOTLIN_STRINGS_FILE_NAME} (Kotlin)")
                        else:
                            self.log_message.emit(f"⚠️ '{key}' no encontrado en {asset_name}/{KOTLIN_STRINGS_FILE_NAME} (Kotlin)")

                except (json.JSONDecodeError, etree.XMLSyntaxError):
                    self.log_message.emit(f"⚠️ Error al leer '{current_path}'. Archivo JSON/XML inválido.")
                except Exception as e:
                    self.log_message.emit(f"❌ Error al procesar '{current_path}': {e}")
            else:
                self.log_message.emit(f"⚠️ Archivo/ubicación no encontrado: {current_path}")

            files_processed_count += 1
            self.progress_updated.emit(files_processed_count)

        self.translation_finished.emit({'key': key, 'undo_data': undo_data, 'platform': platform})

class TranslatorApp(QMainWindow):
    """
    Ventana principal de la aplicación Traductor ARB/Kotlin.
    Gestiona la UI, operaciones de archivos, llamadas a la API, registro, historial y deshacer.
    """
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Traductor ARB/Kotlin - Joss Red")
        self.setMinimumSize(700, 600)
        self.project_path = get_script_dir() # Ruta del proyecto por defecto (directorio del script)
        self.history = []
        self.worker_thread = None
        self.current_platform = "flutter" # Plataforma por defecto

        self.init_ui() 
        self._load_history()
        self._initialize_log_file()
        self._update_project_path_display() # Actualizar el QLineEdit con la ruta inicial

    def init_ui(self):
        """
        Inicializa los elementos de la interfaz de usuario y su diseño.
        """
        main_layout = QVBoxLayout()

        # Selector de plataforma
        platform_layout = QHBoxLayout()
        platform_layout.addWidget(QLabel("Seleccionar Plataforma:"))
        self.platform_selector = QComboBox()
        self.platform_selector.addItem("Flutter (ARB)", "flutter")
        self.platform_selector.addItem("Kotlin (XML)", "kotlin")
        self.platform_selector.currentIndexChanged.connect(self._on_platform_changed)
        platform_layout.addWidget(self.platform_selector)
        platform_layout.addStretch()
        main_layout.addLayout(platform_layout)

        # Selector de carpeta de proyecto
        project_path_layout = QHBoxLayout()
        project_path_layout.addWidget(QLabel("Ruta del Proyecto:"))
        self.project_path_display = QLineEdit(self.project_path)
        self.project_path_display.setReadOnly(True)
        project_path_layout.addWidget(self.project_path_display)
        self.select_folder_btn = QPushButton("Seleccionar Carpeta")
        self.select_folder_btn.clicked.connect(self._select_project_folder)
        project_path_layout.addWidget(self.select_folder_btn)
        main_layout.addLayout(project_path_layout)


        # Campos de entrada
        input_grid_layout = QVBoxLayout()
        input_grid_layout.addWidget(QLabel("Idioma base"))
        self.base_lang_input = QLineEdit()
        self.base_lang_input.setPlaceholderText("Idioma base (ej. 'es', 'en')")
        input_grid_layout.addWidget(self.base_lang_input)

        input_grid_layout.addWidget(QLabel("Texto original"))
        self.text_input = QLineEdit()
        self.text_input.setPlaceholderText("Texto original")
        input_grid_layout.addWidget(self.text_input)

        input_grid_layout.addWidget(QLabel("Nombre de la etiqueta / String"))
        self.key_input = QLineEdit()
        self.key_input.setPlaceholderText("Nombre de la etiqueta (Flutter) / Nombre del string (Kotlin)")
        input_grid_layout.addWidget(self.key_input)

        self.desc_label = QLabel("Descripción (Flutter - opcional)")
        input_grid_layout.addWidget(self.desc_label)
        self.desc_input = QLineEdit()
        self.desc_input.setPlaceholderText("Descripción (opcional)")
        input_grid_layout.addWidget(self.desc_input)

        main_layout.addLayout(input_grid_layout)

        # Botones de acción
        button_layout = QHBoxLayout()
        self.translate_button = QPushButton("Traducir y Agregar")
        self.translate_button.clicked.connect(self._start_translation)
        button_layout.addWidget(self.translate_button)

        self.create_files_btn = QPushButton("Crear Archivos/Carpetas de Idioma")
        self.create_files_btn.clicked.connect(self.create_language_assets)
        button_layout.addWidget(self.create_files_btn)

        self.delete_files_btn = QPushButton("Eliminar Archivos/Carpetas de Idioma")
        self.delete_files_btn.clicked.connect(self.delete_language_assets)
        button_layout.addWidget(self.delete_files_btn)

        self.delete_key_btn = QPushButton("Eliminar Etiqueta/String")
        self.delete_key_btn.clicked.connect(self.delete_key_prompt)
        button_layout.addWidget(self.delete_key_btn)
        main_layout.addLayout(button_layout)
        
        # Botón específico de Flutter Intl Generate
        intl_generate_layout = QHBoxLayout()
        self.flutter_intl_generate_btn = QPushButton("Actualizar Intl de Flutter")
        self.flutter_intl_generate_btn.clicked.connect(self._start_flutter_intl_generate)
        intl_generate_layout.addWidget(self.flutter_intl_generate_btn)
        intl_generate_layout.addStretch()
        main_layout.addLayout(intl_generate_layout)


        # Botones de Historial y Deshacer
        history_undo_layout = QHBoxLayout()
        self.history_btn = QPushButton("Ver Historial")
        self.history_btn.clicked.connect(self.show_history)
        history_undo_layout.addWidget(self.history_btn)

        self.undo_btn = QPushButton("Deshacer Último Cambio")
        self.undo_btn.clicked.connect(self.undo_last_action)
        history_undo_layout.addWidget(self.undo_btn)
        self._update_undo_button_state()
        main_layout.addLayout(history_undo_layout)

        # Barra de progreso
        self.progress_bar = QProgressBar()
        self.progress_bar.setAlignment(Qt.AlignCenter)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setFormat("Progreso: %p%")
        main_layout.addWidget(self.progress_bar)

        main_layout.addSpacing(10)

        # Consola de salida
        main_layout.addWidget(QLabel("Consola de salida"))
        self.output = QTextEdit()
        self.output.setReadOnly(True)
        main_layout.addWidget(self.output)

        container = QWidget()
        container.setLayout(main_layout)
        self.setCentralWidget(container)

        # Establecer el estado inicial de la UI para Flutter
        self._update_ui_for_platform("flutter")

    def _on_platform_changed(self, index):
        """
        Actualiza la plataforma seleccionada y ajusta la UI.
        """
        self.current_platform = self.platform_selector.itemData(index)
        self._update_ui_for_platform(self.current_platform)
        self.log(f"Plataforma cambiada a: {self.current_platform.upper()}")

    def _update_ui_for_platform(self, platform):
        """
        Ajusta la visibilidad y el texto de los elementos de la UI según la plataforma.
        """
        if platform == "flutter":
            self.desc_label.show()
            self.desc_input.show()
            self.key_input.setPlaceholderText("Nombre de la etiqueta")
            self.translate_button.setText("Traducir y Agregar Etiqueta")
            self.create_files_btn.setText("Crear Archivos ARB")
            self.delete_files_btn.setText("Eliminar Archivos ARB")
            self.delete_key_btn.setText("Eliminar Etiqueta de Archivos")
            self.flutter_intl_generate_btn.show() # Mostrar botón de Intl Generate
        elif platform == "kotlin":
            self.desc_label.hide()
            self.desc_input.hide()
            self.key_input.setPlaceholderText("Nombre del string (ej. 'app_name')")
            self.translate_button.setText("Traducir y Agregar String")
            self.create_files_btn.setText("Crear Carpetas y Archivos XML")
            self.delete_files_btn.setText("Eliminar Carpetas y Archivos XML")
            self.delete_key_btn.setText("Eliminar String de Archivos")
            self.flutter_intl_generate_btn.hide() # Ocultar botón de Intl Generate

    def _select_project_folder(self):
        """
        Abre un diálogo para que el usuario seleccione la carpeta del proyecto.
        Actualiza la ruta del proyecto y recarga el historial/log si es necesario.
        """
        dialog = QFileDialog(self)
        dialog.setFileMode(QFileDialog.Directory)
        dialog.setOption(QFileDialog.ShowDirsOnly, True)
        
        selected_dir = dialog.getExistingDirectory(self, "Seleccionar Carpeta del Proyecto", self.project_path)
        
        if selected_dir:
            self.project_path = selected_dir
            self._update_project_path_display()
            self.log(f"Carpeta del proyecto seleccionada: {self.project_path}")
            # Opcional: Recargar historial y re-inicializar log si la ruta del proyecto cambia
            self._load_history()
            self._initialize_log_file()
        else:
            self.log("Selección de carpeta de proyecto cancelada.")

    def _update_project_path_display(self):
        """
        Actualiza el QLineEdit que muestra la ruta del proyecto.
        """
        self.project_path_display.setText(self.project_path)

    def log(self, text):
        """
        Añade un mensaje a la consola de salida.
        """
        self.output.append(text)

    def _initialize_log_file(self):
        """
        Asegura que el archivo de log XLSX exista y tenga los encabezados correctos.
        """
        log_path = os.path.join(self.project_path, LOG_FILE) # Usa self.project_path
        if not os.path.exists(log_path):
            try:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Registro de Traducciones"

                headers = ["Fecha", "Idioma Base", "Texto Original", "Etiqueta", "Descripción", "Acción", "Plataforma"]
                sheet.append(headers)

                for i, header in enumerate(headers):
                    sheet.column_dimensions[get_column_letter(i + 1)].width = 20

                workbook.save(log_path)
                self.log(f"📝 Archivo de log '{LOG_FILE}' creado en {self.project_path}.")
            except Exception as e:
                self.log(f"❌ Error al crear el archivo de log XLSX en {self.project_path}: {e}")

    def _log_action(self, base_lang, original_text, key, desc, action_type, platform):
        """
        Registra una acción en el archivo XLSX.
        """
        log_path = os.path.join(self.project_path, LOG_FILE) # Usa self.project_path
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            workbook = openpyxl.load_workbook(log_path)
            sheet = workbook.active
            sheet.append([timestamp, base_lang, original_text, key, desc, action_type, platform])
            workbook.save(log_path)
            self.log(f"📝 Acción '{action_type}' logeada para '{key}' en {platform.upper()}.")
        except Exception as e:
            self.log(f"❌ Error al escribir en el archivo de log XLSX: {e}")

    def _load_history(self):
        """
        Carga el historial de acciones desde el archivo JSON.
        """
        history_path = os.path.join(self.project_path, HISTORY_FILE) # Usa self.project_path
        if os.path.exists(history_path):
            try:
                with open(history_path, 'r', encoding='utf-8') as f:
                    self.history = json.load(f)
                self.log(f"📚 Historial cargado desde '{HISTORY_FILE}' en {self.project_path}.")
            except json.JSONDecodeError:
                self.log(f"⚠️ Error al leer el historial. El archivo '{HISTORY_FILE}' puede estar corrupto.")
                self.history = []
            except Exception as e:
                self.log(f"❌ Error al cargar el historial: {e}")
                self.history = []
        else:
            self.log(f"ℹ️ No se encontró el archivo de historial '{HISTORY_FILE}' en {self.project_path}. Se creará uno nuevo.")
        self._update_undo_button_state()

    def _save_history(self):
        """
        Guarda el historial de acciones actual en el archivo JSON.
        """
        history_path = os.path.join(self.project_path, HISTORY_FILE) # Usa self.project_path
        try:
            with open(history_path, 'w', encoding='utf-8') as f:
                json.dump(self.history, f, indent=2, ensure_ascii=False)
            self.log(f"💾 Historial guardado en '{HISTORY_FILE}' en {self.project_path}.")
        except Exception as e:
            self.log(f"❌ Error al guardar el historial: {e}")
        self._update_undo_button_state()

    def _add_to_history(self, action_type, data, platform):
        """
        Añade una acción a la lista de historial y la guarda.
        """
        self.history.append({'type': action_type, 'data': data, 'timestamp': datetime.now().isoformat(), 'platform': platform})
        self._save_history()

    def _update_undo_button_state(self):
        """
        Habilita o deshabilita el botón de deshacer según si hay historial.
        """
        self.undo_btn.setEnabled(len(self.history) > 0)

    def create_language_assets(self):
        """
        Crea los archivos/carpetas de idioma según la plataforma seleccionada.
        """
        if self.current_platform == "flutter":
            self._create_flutter_language_files()
        elif self.current_platform == "kotlin":
            self._create_kotlin_language_folders()

    def _create_flutter_language_files(self):
        """
        Crea nuevos archivos ARB para cada idioma si no existen.
        """
        self.log("Iniciando creación de archivos ARB (Flutter)...")
        self.progress_bar.setMaximum(len(FLUTTER_LANGUAGE_FILES))
        for i, file_name in enumerate(FLUTTER_LANGUAGE_FILES):
            path = os.path.join(self.project_path, file_name) # Usa self.project_path
            locale = file_name.split('_')[1].split('.')[0]
            if not os.path.exists(path):
                try:
                    with open(path, "w", encoding="utf-8") as f:
                        json.dump({"@@locale": locale}, f, indent=2, ensure_ascii=False)
                    self.log(f"✅ Archivo creado: {file_name}")
                except Exception as e:
                    self.log(f"❌ Error al crear '{file_name}': {e}")
            else:
                self.log(f"⚠️ Archivo ya existe: {file_name}")
            self.progress_bar.setValue(i + 1)
        self.progress_bar.setValue(0)
        self.log("Creación de archivos ARB (Flutter) finalizada.")

    def _create_kotlin_language_folders(self):
        """
        Crea las carpetas y archivos strings-joss.xml para los idiomas de Kotlin si no existen.
        """
        self.log("Iniciando creación de carpetas y archivos XML (Kotlin)...")
        self.progress_bar.setMaximum(len(KOTLIN_LANGUAGE_FOLDERS))
        for i, folder_name in enumerate(KOTLIN_LANGUAGE_FOLDERS):
            folder_path = os.path.join(self.project_path, folder_name) # Usa self.project_path
            strings_file = os.path.join(folder_path, KOTLIN_STRINGS_FILE_NAME)

            if not os.path.exists(folder_path):
                try:
                    os.makedirs(folder_path)
                    self.log(f"✅ Carpeta creada: {folder_path}")
                except Exception as e:
                    self.log(f"❌ Error al crear carpeta '{folder_path}': {e}")
            else:
                self.log(f"⚠️ Carpeta ya existe: {folder_path}")

            if not os.path.exists(strings_file):
                try:
                    with open(strings_file, "w", encoding="utf-8") as file:
                        file.write("<?xml version='1.0' encoding='UTF-8'?>\n<resources>\n\n</resources>")
                    self.log(f"✅ Archivo creado: {strings_file}")
                except Exception as e:
                    self.log(f"❌ Error al crear archivo '{strings_file}': {e}")
            else:
                self.log(f"⚠️ Archivo ya existente: {strings_file}")
            self.progress_bar.setValue(i + 1)
        self.progress_bar.setValue(0)
        self.log("Creación de carpetas y archivos XML (Kotlin) finalizada.")

    def delete_language_assets(self):
        """
        Elimina los archivos/carpetas de idioma según la plataforma seleccionada.
        """
        msg_box = QMessageBox()
        msg_box.setWindowTitle("Confirmar Eliminación")
        msg_box.setText(f"¿Estás seguro de que quieres eliminar TODOS los archivos/carpetas de idioma para {self.current_platform.upper()}? Esta acción no se puede deshacer fácilmente.")
        msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        msg_box.setDefaultButton(QMessageBox.No)
        reply = msg_box.exec()

        if reply == QMessageBox.No:
            self.log("Operación de eliminación de archivos/carpetas cancelada.")
            return

        if self.current_platform == "flutter":
            self._delete_flutter_language_files()
        elif self.current_platform == "kotlin":
            self._delete_kotlin_language_folders()

    def _delete_flutter_language_files(self):
        """
        Elimina todos los archivos ARB de Flutter.
        """
        self.log("Iniciando eliminación de archivos ARB (Flutter)...")
        self.progress_bar.setMaximum(len(FLUTTER_LANGUAGE_FILES))
        for i, file_name in enumerate(FLUTTER_LANGUAGE_FILES):
            path = os.path.join(self.project_path, file_name) # Usa self.project_path
            if os.path.exists(path):
                try:
                    os.remove(path)
                    self.log(f"🗑️ Archivo eliminado: {file_name}")
                except Exception as e:
                    self.log(f"❌ Error al eliminar '{file_name}': {e}")
            else:
                self.log(f"⚠️ Archivo no encontrado: {file_name}")
            self.progress_bar.setValue(i + 1)
        self.progress_bar.setValue(0)
        self.log("Eliminación de archivos ARB (Flutter) finalizada.")

    def _delete_kotlin_language_folders(self):
        """
        Elimina las carpetas de idioma de Kotlin y sus contenidos.
        """
        self.log("Iniciando eliminación de carpetas y archivos XML (Kotlin)...")
        self.progress_bar.setMaximum(len(KOTLIN_LANGUAGE_FOLDERS))
        for i, folder_name in enumerate(KOTLIN_LANGUAGE_FOLDERS):
            folder_path = os.path.join(self.project_path, folder_name) # Usa self.project_path
            if os.path.exists(folder_path) and os.path.isdir(folder_path):
                try:
                    # Eliminar carpeta y su contenido de forma recursiva
                    for root, dirs, files in os.walk(folder_path, topdown=False):
                        for file in files:
                            os.remove(os.path.join(root, file))
                        for dir in dirs:
                            os.rmdir(os.path.join(root, dir))
                    os.rmdir(folder_path)
                    self.log(f"🗑️ Carpeta eliminada: {folder_path}")
                except Exception as e:
                    self.log(f"❌ Error al eliminar '{folder_path}': {e}")
            else:
                self.log(f"⚠️ Carpeta no encontrada o ya eliminada: {folder_path}")
            self.progress_bar.setValue(i + 1)
        self.progress_bar.setValue(0)
        self.log("Eliminación de carpetas y archivos XML (Kotlin) finalizada.")

    def _start_translation(self):
        """
        Inicia el proceso de traducción en un hilo separado, según la plataforma.
        """
        base_lang = self.base_lang_input.text().strip()
        original_text = self.text_input.text().strip()
        key = self.key_input.text().strip()
        desc = self.desc_input.text().strip() if self.current_platform == "flutter" else "Generado con Joss Red" # Descripción solo para Flutter

        if not all([base_lang, original_text, key]):
            self.log("⚠️ Por favor, completa todos los campos requeridos (Idioma base, Texto original, Nombre de la etiqueta/string).")
            return

        existing_key_files = []
        if self.current_platform == "flutter":
            files_to_check = FLUTTER_LANGUAGE_FILES
        elif self.current_platform == "kotlin":
            files_to_check = [os.path.join(f, KOTLIN_STRINGS_FILE_NAME) for f in KOTLIN_LANGUAGE_FOLDERS]
        else:
            self.log("❌ Plataforma no válida para la verificación de clave.")
            return

        for file_or_folder_path_part in files_to_check:
            full_path = os.path.join(self.project_path, file_or_folder_path_part) # Usa self.project_path
            
            if self.current_platform == "flutter":
                if os.path.exists(full_path):
                    try:
                        with open(full_path, "r", encoding="utf-8") as f:
                            arb_data = json.load(f)
                        if key in arb_data:
                            existing_key_files.append(file_or_folder_path_part)
                    except json.JSONDecodeError:
                        self.log(f"⚠️ Error al leer '{full_path}'. Archivo JSON inválido.")
                    except Exception as e:
                        self.log(f"❌ Error al verificar '{full_path}': {e}")
            elif self.current_platform == "kotlin":
                if os.path.exists(full_path):
                    try:
                        parser = etree.XMLParser(remove_blank_text=True, remove_comments=False)
                        tree = etree.parse(full_path, parser)
                        root = tree.getroot()
                        if any(child.get("name") == key for child in root.xpath("string")):
                            existing_key_files.append(os.path.dirname(file_or_folder_path_part)) 
                    except etree.XMLSyntaxError:
                        self.log(f"⚠️ Error al leer '{full_path}'. Archivo XML inválido.")
                    except Exception as e:
                        self.log(f"❌ Error al verificar '{full_path}': {e}")

        if existing_key_files:
            msg = (f"⚠️ La clave/string '{key}' ya existe en las siguientes ubicaciones y no será sobrescrita:\n"
                   + "\n".join(existing_key_files)
                   + "\n\nLa traducción continuará para las ubicaciones donde la clave/string no existe.")
            self.log(msg)
            QMessageBox.information(self, "Clave/String existente en algunas ubicaciones", msg)

        self._set_ui_enabled(False)
        self.progress_bar.setValue(0)
        self.progress_bar.setFormat("Traduciendo y agregando: %p%")

        self.worker_thread = WorkerThread(
            operation_type="translate_and_add",
            data={
                'base_lang': base_lang,
                'original_text': original_text,
                'key': key,
                'desc': desc,
                'existing_key_files': existing_key_files
            },
            platform=self.current_platform,
            project_path=self.project_path # Pasa la ruta del proyecto al hilo de trabajo
        )
        self.worker_thread.progress_updated.connect(self.progress_bar.setValue)
        self.worker_thread.log_message.connect(self.log)
        self.worker_thread.translation_finished.connect(self._finish_translation_and_add)
        self.worker_thread.error_occurred.connect(self._handle_translation_error)
        self.worker_thread.start()

    def _finish_translation_and_add(self, result_data):
        """
        Se llama cuando el hilo de trabajo de traducción finaliza con éxito.
        Añade traducciones a los archivos ARB/XML, registra la acción y actualiza el historial.
        """
        base_lang = result_data['base_lang']
        original_text = result_data['original_text']
        key = result_data['key']
        desc = result_data['desc']
        translations = result_data['translations']
        existing_key_files = result_data.get('existing_key_files', [])
        platform = result_data['platform']

        undo_data = {
            'base_lang': base_lang,
            'original_text': original_text,
            'key': key,
            'desc': desc,
            'affected_files': {}
        }

        if platform == "flutter":
            target_assets = FLUTTER_LANGUAGE_FILES
        elif platform == "kotlin":
            target_assets = KOTLIN_LANGUAGE_FOLDERS
        else:
            self.log(f"❌ Plataforma desconocida: {platform}")
            self.progress_bar.setValue(0)
            self._set_ui_enabled(True)
            return

        files_processed_count = 0
        total_assets = len(target_assets)
        self.progress_bar.setMaximum(total_assets)
        self.progress_bar.setFormat(f"Guardando traducciones ({platform.upper()}): %p%")

        for i, asset_name in enumerate(target_assets):
            if platform == "flutter":
                current_path = os.path.join(self.project_path, asset_name) # Usa self.project_path
                lang = asset_name.split('_')[1].split('.')[0]
                asset_identifier_for_check = asset_name 
            elif platform == "kotlin":
                current_path = os.path.join(self.project_path, asset_name, KOTLIN_STRINGS_FILE_NAME) # Usa self.project_path
                lang = "en" if asset_name == "values" else asset_name.replace("values-", "")
                asset_identifier_for_check = asset_name 
            
            text = original_text if lang == base_lang else translations.get(lang)
            if not text and '-r' in lang:
                simple_lang = lang.split('-r')[0]
                text = translations.get(simple_lang)
                if text:
                    self.log(f"ℹ️ Usando traducción de '{simple_lang}' para '{lang}'.")
                else:
                    self.log(f"⚠️ Traducción omitida para {lang} (texto no disponible).")
                    files_processed_count += 1
                    self.progress_bar.setValue(files_processed_count)
                    continue
            elif not text:
                self.log(f"⚠️ Traducción omitida para {lang} (texto no disponible).")
                files_processed_count += 1
                self.progress_bar.setValue(files_processed_count)
                continue

            if asset_identifier_for_check in existing_key_files:
                self.log(f"ℹ️ Clave/string '{key}' ya existe en '{asset_identifier_for_check}'. Se omite la adición para esta ubicación.")
                files_processed_count += 1
                self.progress_bar.setValue(files_processed_count)
                continue

            if not os.path.exists(current_path):
                self.log(f"❌ Archivo/ubicación no encontrado: {current_path}. Creando...")
                if platform == "flutter":
                    try:
                        with open(current_path, "w", encoding="utf-8") as f:
                            json.dump({"@@locale": lang}, f, indent=2, ensure_ascii=False)
                    except Exception as e:
                        self.log(f"❌ Error al crear archivo ARB '{current_path}': {e}")
                        files_processed_count += 1
                        self.progress_bar.setValue(files_processed_count)
                        continue
                elif platform == "kotlin":
                    try:
                        os.makedirs(os.path.dirname(current_path), exist_ok=True)
                        with open(current_path, "w", encoding="utf-8") as f:
                            f.write("<?xml version='1.0' encoding='UTF-8'?>\n<resources>\n\n</resources>")
                    except Exception as e:
                        self.log(f"❌ Error al crear archivo XML '{current_path}': {e}")
                        files_processed_count += 1
                        self.progress_bar.setValue(files_processed_count)
                        continue

            try:
                if platform == "flutter":
                    with open(current_path, "r", encoding="utf-8") as f:
                        arb_data = json.load(f)
                    
                    undo_data['affected_files'][asset_name] = {
                        'old_value': arb_data.get(key),
                        'old_desc': arb_data.get(f"@{key}", {}).get("description")
                    }

                    arb_data[key] = text
                    arb_data[f"@{key}"] = {"description": desc}

                    with open(current_path, "w", encoding="utf-8") as f:
                        json.dump(arb_data, f, indent=2, ensure_ascii=False)
                    self.log(f"✅ Añadido '{key}' en {asset_name} (Flutter)")

                elif platform == "kotlin":
                    parser = etree.XMLParser(remove_blank_text=True, remove_comments=False)
                    tree = etree.parse(current_path, parser)
                    root = tree.getroot()

                    existing_string_element = root.xpath(f"string[@name='{key}']")
                    if existing_string_element:
                        undo_data['affected_files'][asset_name] = {
                            'old_value': existing_string_element[0].text,
                            'old_desc': None
                        }
                    else:
                        undo_data['affected_files'][asset_name] = {'old_value': None, 'old_desc': None}

                    if not existing_string_element:
                        new_string = etree.Element("string", name=key)
                        new_string.text = text
                        root.append(new_string)

                        formatted_xml = etree.tostring(
                            root,
                            encoding="utf-8",
                            xml_declaration=True,
                            pretty_print=True
                        ).decode("utf-8")
                        if not formatted_xml.endswith("\n</resources>\n"):
                            formatted_xml = formatted_xml.replace("</resources>", "\n</resources>")
                        
                        with open(current_path, "w", encoding="utf-8") as f:
                            f.write(formatted_xml)
                        self.log(f"✅ Añadido '{key}' en {asset_name}/{KOTLIN_STRINGS_FILE_NAME} (Kotlin)")
                    else:
                        self.log(f"⚠️ String '{key}' ya existe en {asset_name}/{KOTLIN_STRINGS_FILE_NAME}. Se omite la adición.")

            except json.JSONDecodeError:
                self.log(f"⚠️ Error al leer '{current_path}'. Archivo JSON/XML inválido.")
            except Exception as e:
                self.log(f"❌ Error al procesar '{current_path}': {e}")

            files_processed_count += 1
            self.progress_bar.setValue(files_processed_count)

        self._add_to_history('add_key', undo_data, platform)
        self._log_action(base_lang, original_text, key, desc, "Traducción y Adición", platform)
        self.log(f"Traducción y adición de etiqueta/string finalizada para {platform.upper()}.")
        self.progress_bar.setValue(0)
        self._set_ui_enabled(True)

    def _handle_translation_error(self, message):
        """
        Maneja los errores reportados por el hilo de trabajo de traducción.
        """
        self.log(message)
        QMessageBox.critical(self, "Error de Traducción", message)
        self.progress_bar.setValue(0)
        self._set_ui_enabled(True)

    def _start_flutter_intl_generate(self):
        """
        Inicia la ejecución del comando 'dart run intl_utils:generate' en un hilo de trabajo.
        """
        self.log("Iniciando comando 'dart run intl_utils:generate'...")
        self._set_ui_enabled(False)
        self.progress_bar.setValue(0)
        self.progress_bar.setFormat("Ejecutando comando: %p%")

        self.worker_thread = WorkerThread(
            operation_type="run_flutter_intl_generate",
            project_path=self.project_path
        )
        self.worker_thread.progress_updated.connect(self.progress_bar.setValue)
        self.worker_thread.log_message.connect(self.log)
        self.worker_thread.command_output.connect(self.log)
        self.worker_thread.translation_finished.connect(self._finish_command_execution)
        self.worker_thread.error_occurred.connect(self._handle_command_error)
        self.worker_thread.start()

    def _finish_command_execution(self):
        """
        Se llama cuando el hilo de ejecución de comando finaliza.
        """
        self.log("Ejecución de comando finalizada.")
        self.progress_bar.setValue(0)
        self._set_ui_enabled(True)

    def _handle_command_error(self, message):
        """
        Maneja los errores reportados por el hilo de ejecución de comando.
        """
        self.log(message)
        QMessageBox.critical(self, "Error de Comando", message)
        self.progress_bar.setValue(0)
        self._set_ui_enabled(True)

    def delete_key_prompt(self):
        """
        Pide al usuario una clave/string para eliminar y luego llama al método de eliminación.
        """
        key, ok = QInputDialog.getText(self, "Eliminar clave/string", "Nombre de la etiqueta/string a eliminar:")
        if ok and key:
            self.log(f"Iniciando eliminación de clave/string '{key.strip()}' en plataforma {self.current_platform.upper()}...")
            self._start_delete_key(key.strip())

    def _start_delete_key(self, key):
        """
        Inicia el proceso de eliminación de clave/string en un hilo separado.
        """
        undo_data = {'key': key, 'deleted_content_per_file': {}}
        
        if self.current_platform == "flutter":
            assets_to_check = FLUTTER_LANGUAGE_FILES
        elif self.current_platform == "kotlin":
            assets_to_check = KOTLIN_LANGUAGE_FOLDERS
        else:
            self.log("❌ Plataforma no válida para la eliminación de clave/string.")
            return

        for asset_name in assets_to_check:
            if self.current_platform == "flutter":
                current_path = os.path.join(self.project_path, asset_name)
            elif self.current_platform == "kotlin":
                current_path = os.path.join(self.project_path, asset_name, KOTLIN_STRINGS_FILE_NAME)

            if os.path.exists(current_path):
                try:
                    if self.current_platform == "flutter":
                        with open(current_path, "r", encoding="utf-8") as f:
                            arb_data = json.load(f)
                        if key in arb_data:
                            undo_data['deleted_content_per_file'][asset_name] = {
                                'value': arb_data.get(key),
                                'description': arb_data.get(f"@{key}", {}).get("description")
                            }
                    elif self.current_platform == "kotlin":
                        parser = etree.XMLParser(remove_blank_text=True, remove_comments=False)
                        tree = etree.parse(current_path, parser)
                        root = tree.getroot()
                        existing_string_element = root.xpath(f"string[@name='{key}']")
                        if existing_string_element:
                            undo_data['deleted_content_per_file'][asset_name] = {
                                'value': existing_string_element[0].text,
                                'description': None
                            }
                except (json.JSONDecodeError, etree.XMLSyntaxError):
                    self.log(f"⚠️ Error al leer '{current_path}'. Archivo inválido. No se puede guardar para deshacer.")
                except Exception as e:
                    self.log(f"❌ Error al acceder a '{current_path}': {e}. No se puede guardar para deshacer.")

        self._set_ui_enabled(False)
        self.progress_bar.setValue(0)
        self.progress_bar.setFormat("Eliminando clave/string: %p%")

        self.worker_thread = WorkerThread(
            operation_type="delete_key",
            data={'key': key, 'undo_data': undo_data},
            platform=self.current_platform,
            project_path=self.project_path
        )
        self.worker_thread.progress_updated.connect(self.progress_bar.setValue)
        self.worker_thread.log_message.connect(self.log)
        self.worker_thread.translation_finished.connect(self._finish_delete_key)
        self.worker_thread.error_occurred.connect(self._handle_deletion_error)
        self.worker_thread.start()

    def _perform_delete_key(self):
        """
        Método del hilo de trabajo para realizar la eliminación real de una clave/string.
        """
        key = self.data['key']
        undo_data = self.data['undo_data']
        platform = self.platform

        if platform == "flutter":
            target_assets = FLUTTER_LANGUAGE_FILES
        elif platform == "kotlin":
            target_assets = KOTLIN_LANGUAGE_FOLDERS
        else:
            self.log_message.emit(f"❌ Plataforma desconocida para la eliminación: {platform}")
            self.error_occurred.emit(f"Plataforma desconocida: {platform}")
            return

        files_processed_count = 0
        total_assets = len(target_assets)
        self.progress_updated.emit(0)

        for i, asset_name in enumerate(target_assets):
            if platform == "flutter":
                current_path = os.path.join(self.project_path, asset_name)
            elif platform == "kotlin":
                current_path = os.path.join(self.project_path, asset_name, KOTLIN_STRINGS_FILE_NAME)

            if os.path.exists(current_path):
                try:
                    if platform == "flutter":
                        with open(current_path, "r", encoding="utf-8") as f:
                            data = json.load(f)

                        if key in data:
                            del data[key]
                            data.pop(f"@{key}", None)

                            with open(current_path, "w", encoding="utf-8") as f:
                                json.dump(data, f, indent=2, ensure_ascii=False)
                            self.log_message.emit(f"🗑️ '{key}' eliminado de {asset_name} (Flutter)")
                        else:
                            self.log_message.emit(f"⚠️ '{key}' no encontrado en {asset_name} (Flutter)")
                    elif platform == "kotlin":
                        parser = etree.XMLParser(remove_blank_text=True, remove_comments=False)
                        tree = etree.parse(current_path, parser)
                        root = tree.getroot()

                        target_string = root.xpath(f"string[@name='{key}']")
                        if target_string:
                            root.remove(target_string[0])
                            formatted_xml = etree.tostring(
                                root,
                                encoding="utf-8",
                                xml_declaration=True,
                                pretty_print=True
                            ).decode("utf-8")
                            if not formatted_xml.endswith("\n</resources>\n"):
                                formatted_xml = formatted_xml.replace("</resources>", "\n</resources>")
                            with open(current_path, "w", encoding="utf-8") as f:
                                f.write(formatted_xml)
                            self.log_message.emit(f"🗑️ '{key}' eliminado de {asset_name}/{KOTLIN_STRINGS_FILE_NAME} (Kotlin)")
                        else:
                            self.log_message.emit(f"⚠️ '{key}' no encontrado en {asset_name}/{KOTLIN_STRINGS_FILE_NAME} (Kotlin)")

                except (json.JSONDecodeError, etree.XMLSyntaxError):
                    self.log_message.emit(f"⚠️ Error al leer '{current_path}'. Archivo JSON/XML inválido.")
                except Exception as e:
                    self.log_message.emit(f"❌ Error al procesar '{current_path}': {e}")
            else:
                self.log_message.emit(f"⚠️ Archivo/ubicación no encontrado: {current_path}")

            files_processed_count += 1
            self.progress_updated.emit(files_processed_count)

        self.translation_finished.emit({'key': key, 'undo_data': undo_data, 'platform': platform})

    def _finish_delete_key(self, result_data):
        """
        Se llama cuando el hilo de trabajo de eliminación de clave/string finaliza con éxito.
        Registra la acción y actualiza el historial.
        """
        key = result_data['key']
        undo_data = result_data['undo_data']
        platform = result_data['platform']

        self._add_to_history('delete_key', undo_data, platform)
        self._log_action("", "", key, "", "Eliminación de Etiqueta/String", platform)
        self.log(f"Eliminación de clave/string '{key}' finalizada para {platform.upper()}.")
        self.progress_bar.setValue(0)
        self._set_ui_enabled(True)

    def _handle_deletion_error(self, message):
        """
        Maneja los errores reportados por el hilo de trabajo de eliminación.
        """
        self.log(message)
        QMessageBox.critical(self, "Error de Eliminación", message)
        self.progress_bar.setValue(0)
        self._set_ui_enabled(True)

    def _set_ui_enabled(self, enabled):
        """
        Habilita o deshabilita todos los botones e inputs principales de la UI.
        """
        self.translate_button.setEnabled(enabled)
        self.create_files_btn.setEnabled(enabled)
        self.delete_files_btn.setEnabled(enabled)
        self.delete_key_btn.setEnabled(enabled)
        self.history_btn.setEnabled(enabled)
        self.base_lang_input.setEnabled(enabled)
        self.text_input.setEnabled(enabled)
        self.key_input.setEnabled(enabled)
        self.desc_input.setEnabled(enabled)
        self.platform_selector.setEnabled(enabled)
        self.select_folder_btn.setEnabled(enabled) # Habilitar/deshabilitar el botón de selección de carpeta
        
        # El botón de Flutter Intl Generate solo se habilita si la plataforma es Flutter
        self.flutter_intl_generate_btn.setEnabled(enabled and self.current_platform == "flutter")
        
        self._update_undo_button_state()

    def show_history(self):
        """
        Muestra el historial de traducciones en una nueva ventana de diálogo.
        """
        history_dialog = QDialog(self)
        history_dialog.setWindowTitle("Historial de Acciones")
        history_dialog.setMinimumSize(600, 400)
        layout = QVBoxLayout()

        history_list_widget = QListWidget()
        if not self.history:
            history_list_widget.addItem("No hay historial disponible.")
        else:
            for entry in reversed(self.history):
                action_type = entry.get('type', 'Desconocido')
                timestamp = datetime.fromisoformat(entry.get('timestamp')).strftime("%Y-%m-%d %H:%M:%S")
                data = entry.get('data', {})
                platform_hist = entry.get('platform', 'Desconocida')

                display_text = f"[{timestamp}] Plataforma: {platform_hist.upper()} - Tipo: {action_type.replace('_', ' ').title()}"

                if action_type == 'add_key':
                    key = data.get('key', 'N/A')
                    base_lang = data.get('base_lang', 'N/A')
                    original_text = data.get('original_text', 'N/A')
                    display_text += f" - Clave/String: '{key}', Idioma Base: '{base_lang}', Texto: '{original_text}'"
                elif action_type == 'delete_key':
                    key = data.get('key', 'N/A')
                    display_text += f" - Clave/String: '{key}'"

                history_list_widget.addItem(display_text)

        layout.addWidget(history_list_widget)
        history_dialog.setLayout(layout)
        history_dialog.exec()

    def undo_last_action(self):
        """
        Intenta deshacer la última acción registrada en el historial.
        """
        if not self.history:
            self.log("⚠️ No hay acciones en el historial para deshacer.")
            return

        last_action = self.history[-1]
        action_type = last_action['type']
        data = last_action['data']
        platform = last_action['platform']

        msg_box = QMessageBox()
        msg_box.setWindowTitle("Confirmar Deshacer")
        msg_box.setText(f"¿Estás seguro de que quieres deshacer la última acción ({action_type.replace('_', ' ').title()}) para {platform.upper()}?")
        msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        msg_box.setDefaultButton(QMessageBox.No)
        reply = msg_box.exec()

        if reply == QMessageBox.No:
            self.log("Operación de deshacer cancelada.")
            return

        self.log(f"Intentando deshacer la acción: {action_type} en {platform.upper()}...")
        self._set_ui_enabled(False)

        try:
            if action_type == 'add_key':
                key_to_delete = data['key']
                self.log(f"Deshaciendo: Eliminando clave/string '{key_to_delete}' de {platform.upper()}...")
                self._undo_delete_key(key_to_delete, platform)
                self.history.pop()
                self._save_history()
                self.log(f"✅ Acción 'add_key' deshecha para '{key_to_delete}' en {platform.upper()}.")

            elif action_type == 'delete_key':
                key_to_restore = data['key']
                deleted_content = data['deleted_content_per_file']
                self.log(f"Deshaciendo: Restaurando clave/string '{key_to_restore}' en {platform.upper()}...")
                self._undo_add_key(key_to_restore, deleted_content, platform)
                self.history.pop()
                self._save_history()
                self.log(f"✅ Acción 'delete_key' deshecha para '{key_to_restore}' en {platform.upper()}.")

            else:
                self.log(f"⚠️ Tipo de acción '{action_type}' no soportado para deshacer.")
                QMessageBox.warning(self, "Deshacer", "Tipo de acción no soportado para deshacer.")

        except Exception as e:
            self.log(f"❌ Error al deshacer la acción: {e}")
            QMessageBox.critical(self, "Error al Deshacer", f"Ocurrió un error al intentar deshacer: {e}")
        finally:
            self._set_ui_enabled(True)

    def _undo_delete_key(self, key_to_delete, platform):
        """
        Método interno para eliminar una clave/string de todos los archivos/ubicaciones,
        usado para deshacer una acción 'add_key'.
        """
        if platform == "flutter":
            target_assets = FLUTTER_LANGUAGE_FILES
        elif platform == "kotlin":
            target_assets = KOTLIN_LANGUAGE_FOLDERS
        else:
            self.log(f"❌ Plataforma desconocida para deshacer eliminación: {platform}")
            return

        self.progress_bar.setMaximum(len(target_assets))
        self.progress_bar.setFormat(f"Deshaciendo eliminación ({platform.upper()}): %p%")
        for i, asset_name in enumerate(target_assets):
            if platform == "flutter":
                current_path = os.path.join(self.project_path, asset_name)
            elif platform == "kotlin":
                current_path = os.path.join(self.project_path, asset_name, KOTLIN_STRINGS_FILE_NAME)

            if os.path.exists(current_path):
                try:
                    if platform == "flutter":
                        with open(current_path, "r", encoding="utf-8") as f:
                            data = json.load(f)

                        if key_to_delete in data:
                            del data[key_to_delete]
                            data.pop(f"@{key_to_delete}", None)

                            with open(current_path, "w", encoding="utf-8") as f:
                                json.dump(data, f, indent=2, ensure_ascii=False)
                            self.log(f"🗑️ '{key_to_delete}' eliminado de {asset_name} (deshecho Flutter).")
                        else:
                            self.log(f"⚠️ '{key_to_delete}' no encontrado en {asset_name} (para deshacer Flutter).")
                    elif platform == "kotlin":
                        parser = etree.XMLParser(remove_blank_text=True, remove_comments=False)
                        tree = etree.parse(current_path, parser)
                        root = tree.getroot()

                        target_string = root.xpath(f"string[@name='{key_to_delete}']")
                        if target_string:
                            root.remove(target_string[0])
                            formatted_xml = etree.tostring(
                                root,
                                encoding="utf-8",
                                xml_declaration=True,
                                pretty_print=True
                            ).decode("utf-8")
                            if not formatted_xml.endswith("\n</resources>\n"):
                                formatted_xml = formatted_xml.replace("</resources>", "\n</resources>")
                            with open(current_path, "w", encoding="utf-8") as f:
                                f.write(formatted_xml)
                            self.log(f"🗑️ '{key_to_delete}' eliminado de {asset_name}/{KOTLIN_STRINGS_FILE_NAME} (deshecho Kotlin).")
                        else:
                            self.log(f"⚠️ '{key_to_delete}' no encontrado en {asset_name}/{KOTLIN_STRINGS_FILE_NAME} (para deshacer Kotlin).")

                except (json.JSONDecodeError, etree.XMLSyntaxError):
                    self.log(f"⚠️ Error al leer '{current_path}'. Archivo JSON/XML inválido (para deshacer).")
                except Exception as e:
                    self.log(f"❌ Error al procesar '{current_path}' (para deshacer): {e}")
            else:
                self.log(f"⚠️ Archivo/ubicación no encontrado: {current_path} (para deshacer).")
            self.progress_bar.setValue(i + 1)
        self.progress_bar.setValue(0)

    def _undo_add_key(self, key_to_restore, deleted_content_per_file, platform):
        """
        Método interno para volver a añadir una clave/string con su contenido anterior,
        usado para deshacer una acción 'delete_key'.
        """
        if platform == "flutter":
            target_assets = FLUTTER_LANGUAGE_FILES
        elif platform == "kotlin":
            target_assets = KOTLIN_LANGUAGE_FOLDERS
        else:
            self.log(f"❌ Plataforma desconocida para deshacer adición: {platform}")
            return

        self.progress_bar.setMaximum(len(target_assets))
        self.progress_bar.setFormat(f"Deshaciendo adición ({platform.upper()}): %p%")
        for i, asset_name in enumerate(target_assets):
            if platform == "flutter":
                current_path = os.path.join(self.project_path, asset_name)
            elif platform == "kotlin":
                current_path = os.path.join(self.project_path, asset_name, KOTLIN_STRINGS_FILE_NAME)

            content_to_restore = deleted_content_per_file.get(asset_name)

            if not content_to_restore or content_to_restore['value'] is None:
                self.log(f"⚠️ No hay contenido para restaurar '{key_to_restore}' en {asset_name}.")
                self.progress_bar.setValue(i + 1)
                continue

            try:
                if platform == "flutter":
                    arb_data = {"@@locale": asset_name.split('_')[1].split('.')[0]}
                    if os.path.exists(current_path):
                        with open(current_path, "r", encoding="utf-8") as f:
                            arb_data = json.load(f)

                    arb_data[key_to_restore] = content_to_restore['value']
                    if content_to_restore['description']:
                        arb_data[f"@{key_to_restore}"] = {"description": content_to_restore['description']}
                    else:
                        arb_data.pop(f"@{key_to_restore}", None)

                    with open(current_path, "w", encoding="utf-8") as f:
                        json.dump(arb_data, f, indent=2, ensure_ascii=False)
                    self.log(f"✅ Restaurado '{key_to_restore}' en {asset_name} (deshecho Flutter).")

                elif platform == "kotlin":
                    parser = etree.XMLParser(remove_blank_text=True, remove_comments=False)
                    if not os.path.exists(current_path):
                        os.makedirs(os.path.dirname(current_path), exist_ok=True)
                        with open(current_path, "w", encoding="utf-8") as f:
                            f.write("<?xml version='1.0' encoding='UTF-8'?>\n<resources>\n\n</resources>")
                        tree = etree.parse(current_path, parser)
                        root = tree.getroot()
                    else:
                        tree = etree.parse(current_path, parser)
                        root = tree.getroot()

                    existing_string_element = root.xpath(f"string[@name='{key_to_restore}']")
                    if not existing_string_element:
                        new_string = etree.Element("string", name=key_to_restore)
                        new_string.text = content_to_restore['value']
                        root.append(new_string)

                        formatted_xml = etree.tostring(
                            root,
                            encoding="utf-8",
                            xml_declaration=True,
                            pretty_print=True
                        ).decode("utf-8")
                        if not formatted_xml.endswith("\n</resources>\n"):
                            formatted_xml = formatted_xml.replace("</resources>", "\n</resources>")
                        with open(current_path, "w", encoding="utf-8") as f:
                            f.write(formatted_xml)
                        self.log(f"✅ Restaurado '{key_to_restore}' en {asset_name}/{KOTLIN_STRINGS_FILE_NAME} (deshecho Kotlin).")
                    else:
                        self.log(f"⚠️ El string '{key_to_restore}' ya existe en {asset_name}/{KOTLIN_STRINGS_FILE_NAME}. No se restauró para evitar duplicados.")


            except (json.JSONDecodeError, etree.XMLSyntaxError):
                self.log(f"⚠️ Error al leer '{current_path}'. Archivo JSON/XML inválido (para deshacer).")
            except Exception as e:
                self.log(f"❌ Error al restaurar '{key_to_restore}' en '{current_path}' (deshacer): {e}")
            self.progress_bar.setValue(i + 1)
        self.progress_bar.setValue(0)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = TranslatorApp()
    window.show()
    sys.exit(app.exec())
